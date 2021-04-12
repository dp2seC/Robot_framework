import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
import os


def Parse_Dict_With_Col(master_result_file, crf_version, study_name, table_name, list_of_dict, col_list):
    ''' from list_of_dict extracts only columns in col_list and return those value as a list of set'''
    try:
        to_df = pd.DataFrame(list_of_dict)
        to_df_with_req_col = to_df[col_list].copy()
        records = to_df_with_req_col.to_records(index=False)
        records = list(records)
        print('*DEBUG* ==Records==', records)
        if len(col_list) == 1:
            value = 'magic'
            records = [tuple([value]+list(each)) for each in records]
            print('*DEBUG* ==Records after adding magic==', records)
            return records
        return records
    except Exception as e:
        Result_to_Master(master_result_file=master_result_file, crf_version=crf_version,
                         study_name=study_name, table_name=table_name, exception=e)


def Prepare_Source_Target_Validate(mapping_rslt, lookup_rslt, target_rslt, merger_col, 
                                   data_structure, sort_val, result_dir, diff_csv_dir, table_name, master_result_file, crf_version, study_name, log_level):
    ''' Gets 3 dicts, convert to df, inner merge first 2 on merger_col, sort the df, compare it with the 
    third df. If there is a differece, it will put that to a csv with count, if not exit the function
    only printing the count'''
    try:
        mapping_df = pd.DataFrame(mapping_rslt)
        lookup_df = pd.DataFrame(lookup_rslt)
        target_df = pd.DataFrame(target_rslt)
        merged_df = pd.merge(left=mapping_df, right=lookup_df, how='left', left_on=merger_col, right_on=merger_col)
        source_data = merged_df[data_structure]
        source_data = source_data.sort_values(sort_val)
        if log_level == 'DEBUG':
            source_data.to_csv('mergeresult_'+table_name+'.csv')
        target_data = target_df[data_structure]
        target_data = target_data.sort_values(sort_val)
        filename = ''
        #Below code for finding difference
        diff_df = pd.concat([source_data, target_data])
        diff_df = diff_df.reset_index(drop=True)
        diff_df_gpby = diff_df.groupby(list(diff_df.columns))
        idx = [x[0] for x in diff_df_gpby.groups.values() if len(x) == 1]
        diff_df = diff_df.reindex(idx)
        if diff_df.empty:
            result = {"result": "Both source and target matches for table", "status":"PASS"}
        else:
            filename = table_name+'.csv'
            diff_df = diff_df.sort_values(sort_val)
            if not os.path.exists(diff_csv_dir):
                os.makedirs(diff_csv_dir)
            diff_df.to_csv(diff_csv_dir+filename, index=False)
            result = {"result": "Difference in table found", "status": "FAIL"}
        len_source = len(source_data.index)
        len_target = len(target_data.index)
        Result_to_Master(master_result_file, crf_version, study_name, table_name,len_source, len_target, filename)
        return result
    except Exception as e:
        Result_to_Master(master_result_file=master_result_file, crf_version=crf_version,
                         study_name=study_name, table_name=table_name, exception=e)
        return {"result": "There is an exception", "exception": e, "status": "FAIL"}


def append_df_to_excel(filename,df,sheet_name='Sheet1',startrow=None,truncate_sheet=False,**to_excel_kwargs):
    ''' except df, prints it to the xlsx file based on the columns. If file not exists it creates one.
    '''
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow,
                columns=['Date', 'CRF Version', 'Study Name', 'TableName', 'Source Count', 'Target Count', 'File', 'Exception'], **to_excel_kwargs)
    writer.save()


def Result_to_Master(master_result_file, crf_version, study_name, table_name, len_source=-1, len_target=-1, filename='', exception=''):
    ''' functions calls append_df_to_excel with an exception messagge, created to handle exceptions.
    '''
    date_time = datetime.datetime.today().strftime('%Y-%m-%d-%H:%M:%S')
    if filename != '':
        exception = 'Data Mismatch in Source and Target'
    if len_source != len_target:
        exception = 'Count Mismatch in Source and Target'
    to_master_result = [{'Date': date_time, 'CRF Version': crf_version, 'Study Name': study_name, 'TableName': table_name,
                         'Source Count': len_source, 'Target Count': len_target, 'File': filename, 'Exception': exception}]

    to_master_result_df = pd.DataFrame(to_master_result)
    append_df_to_excel(master_result_file, to_master_result_df,index=False, header=None)
